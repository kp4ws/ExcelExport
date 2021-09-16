import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import json

class ExtractExcelInsert:
    FILE_NAME = '.\excel_config.json'

    #Constructor
    def __init__(self):
        #Get data from JSON config file
        self._config = self.retrieveData()
        
        #Initialize variables
        self._wb_name = ''
        self._directory = ''
        self._rows = ''
        self._columns = ''
        self._col_start = ''

        if(str(self._config) != '{}' and input('Use recent config? (Enter = yes, Value = no) ') == ''):
            self._wb_name = self._config['workbook']
            self._directory = self._config['destination']
            self._rows = self._config['rows']
            self._columns = self._config['columns']
            self._col_start = self._config['start']
        else:
            self.userInput()
            self.saveConfig()
        
        self._wb = load_workbook(self._wb_name, data_only=True)
        self._ws = self._wb.active
        self.exportData()

    #Get data to export
    def retrieveData(self):
        with open(self.FILE_NAME) as f:
            return json.load(f)

    def userInput(self):
        self._wb_name = input('Enter absolute path for Workbook "C:\\directory name\\<file name>.xlsx" ')
        self._directory = input('Enter absolute path for Export Location "C:\\directory name\\<file name>.sql" ')
        self._rows = int(input('Rows: '))
        self._columns = int(input('Columns: '))
        self._col_start = int(input('Enter column start '))

    def saveConfig(self):
        new_config = {
            "workbook" : self._wb_name,
            "destination" : self._directory,
            "rows" : self._rows,
            "columns" : self._columns,
            "start" : self._col_start
        }

        json_obj = json.dumps(new_config, indent = 6)
        with open(self.FILE_NAME, "w") as f:
            f.write(json_obj)

    #Export data from excel to sql file
    def exportData(self):
        with open (self._directory, 'w') as f:
            for i in range(2, self._rows+1):
                f.write('INSERT INTO table_name VALUES(')
                data_str = ''

                for j in range(self._col_start, self._columns + self._col_start):
                    
                    if(self._ws.cell(row=i,column=j).value is None or self._ws.cell(row=i,column=j).value == 'NULL'):
                        data_str += 'NULL,'

                    elif (isinstance(self._ws.cell(row=i,column=j).value, int)):
                        data_str += str(self._ws.cell(row=i, column=j).value).strip() + ","

                    else:
                        data_str += "'"+ str(self._ws.cell(row=i, column=j).value).strip() + "',"

                else:
                    f.write(data_str[:-1] + ')\n')
                
        print('Complete')


#Main Method
def main():
    ExtractExcelInsert()

if __name__=='__main__':    
    main()
