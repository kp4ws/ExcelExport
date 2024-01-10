import openpyxl
from openpyxl import load_workbook

def generate_insert_sql_commands(data):
    source_file = data["source_file"]
    row_start = data["row_start"]
    rows = data["rows"]
    column_start = data["column_start"]
    columns = data["columns"]
    table_name = data["table_name"]

    sql_commands = []

    workbook = load_workbook(source_file, data_only=True)
    worksheet = workbook.active
    
    for i in range(row_start, row_start + rows):
        values = []

        for j in range(column_start, column_start + columns):
            cell_value = worksheet.cell(row=i, column=j).value

            if cell_value is None or cell_value == 'NULL':
                values.append('NULL')
            elif isinstance(cell_value, int):
                values.append(str(cell_value).strip())
            else:
                values.append("'" + str(cell_value).strip().replace("'", "''") + "'")

        command = ''
        if table_name == '':
            command = f'INSERT INTO {table_name} VALUES(' + ','.join(values) + ")"
        else:
            command = 'INSERT INTO VALUES(' + ','.join(values) + ")"
        
        sql_commands.append(command)

    return sql_commands
