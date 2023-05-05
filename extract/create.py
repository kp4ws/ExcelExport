import openpyxl
from openpyxl import load_workbook
from datetime import datetime

class ExtractExcelCreate:
    def __init__(self, wb_name, file, rows, columns, delimiter):
        self._wb_name = wb_name
        self._file = file
        self._rows = rows
        self._columns = columns
        self._delimiter = delimiter

        self._wb = load_workbook(self._wb_name, data_only=True)
        self._ws = self._wb.active
        self.exportData()

    def exportData(self):
        with open (self._file, 'w') as f:

            #initial CREATE TABLE statement (currently being hardcoded)
            f.write('CREATE TABLE ' + str(self._ws.cell(row=2, column=1).value).strip() + '(\n')
            for i in range(2, self._rows+1):
                
                #Reset data string per iteration of the loop
                data_str = ''
                #Check for delimiter on next line/row
                delim_check = str(self._ws.cell(row=i+1, column=2).value).strip()
                
                for j in range(2, self._columns + 1):

                    #Skip delimiter line
                    if(self._ws.cell(row=i, column=j).value == self._delimiter):
                        break

                    #If constraint column empty, don't append anything to this line
                    elif(self._ws.cell(row=i,column=j).value is not None):
                        data_str += str(self._ws.cell(row=i, column=j).value).strip() + ' '

                else:
                    #Check for when next CREATE TABLE should happen.
                    #NOTE, We need to check this in advance, otherwise a comma will be added to the last table column which is incorrect
                    if(delim_check == self._delimiter):
                        
                        #If table name column empty then append table_name
                        if(self._ws.cell(row=i+1, column=1).value is not None):
                            f.write(data_str + '\n);\n\nCREATE TABLE ' + str(self._ws.cell(row=i+1, column=1).value).strip() + ' (\n')
                        else:
                            f.write(data_str + '\n);\n\nCREATE TABLE table_name (\n')

                    else:
                        f.write(data_str[:-1] + ',\n')
            else:
                f.write(');')