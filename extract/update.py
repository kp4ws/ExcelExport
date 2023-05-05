from turtle import update
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

class ExtractExcelUpdate:
    def __init__(self, wb_name, file, rows, columns, row_start, col_start):
        self._wb_name = wb_name
        self._file = file
        self._rows = rows
        self._columns = columns
        self._row_start = row_start
        self._col_start = col_start

        self._wb = load_workbook(self._wb_name, data_only=True)
        self._ws = self._wb.active
        self.exportData()

    def exportData(self):
        with open (self._file, 'w') as f:

            for i in range(self._row_start, self._rows + self._row_start):
                
                if(self._ws.cell(row=i,column=2).value is None):
                    continue

                f.write('UPDATE table_name\n')
                for j in range(self._col_start, self._columns + self._col_start):


                    originalVal = str(self._ws.cell(row=i, column=j).value).strip()
                    updatedVal = str(self._ws.cell(row=i, column=j + 1).value).strip()
                    
                    f.write("SET column" + str(j) + " = '" + updatedVal + "'\n")
                    # if(self._ws.cell(row=i,column=j).value is None or self._ws.cell(row=i,column=j).value == 'NULL'):
                    #     f.write('NULL\n')

                    # elif (isinstance(self._ws.cell(row=i,column=j).value, int)):
                    #     f.write(str(self._ws.cell(row=i, column=j).value).strip() + '\n')

                    # else:
                    #     f.write("'"+ str(self._ws.cell(row=i, column=j).value).strip() + "'\n")

                    
                    f.write("WHERE column" + str(j) + " = '" + originalVal + "'")
                        
                else:
                    f.write('\n\n')


#This file currently updates 1 row using two columns of data
#First column is the original values currently in the database
#Second column is the updated value