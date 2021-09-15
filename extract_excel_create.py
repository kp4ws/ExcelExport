import openpyxl
from openpyxl import load_workbook
from datetime import datetime

class ExtractExcelUpdate:

    #Constructor
    def __init__(self):

        #Get user input
        print('Enter absolute path for Workbook "C:\\directory name\\<file name>.xlsx"')
        self._wb_name = input()
        #self._wb = load_workbook('.\\'+self._wb_name+'.xlsx', data_only=True)
        self._wb = load_workbook(self._wb_name, data_only=True)
        self._ws = self._wb.active

        print('Enter absolute path for Export Location "C:\\directory name\\<file name>.sql"')
        self._directory = input()
        
        print('Rows:')
        self._rows = int(input())
        #print('Columns:')
        #self._columns = int(input())
        self._columns = 4

        #table delimiter
        print('Enter delimiter')
        self._delimiter = input()

    	#Primary method for exporting data
        self.exportData()

    #Export data from excel to sql file
    def exportData(self):
        with open (self._directory, 'w') as f:

            #initial CREATE TABLE statement (currently being hardcoded)
            f.write('CREATE TABLE ' + str(self._ws.cell(row=2, column=1).value).strip() + '(\n')
            for i in range(2, self._rows+1):
                
                #Reset data string per iteration of the loop
                data_str = ''
                #Check for delimiter on next line/row
                delim_check = str(self._ws.cell(row=i+1, column=2).value).strip()
                
                for j in range(2, self._columns + 1):

                    #Skip delimiter line
                    if(self._ws.cell(row=i, column=j).value == self._delimiter):
                        break

                    #If constraint column empty, don't append anything to this line
                    elif(self._ws.cell(row=i,column=j).value is not None):
                        data_str += str(self._ws.cell(row=i, column=j).value).strip() + ' '

                else:
                    #Check for when next CREATE TABLE should happen.
                    #NOTE, We need to check this in advance, otherwise a comma will be added to the last table column which is incorrect
                    if(delim_check == self._delimiter):
                        if(self._ws.cell(row=i+1, column=1).value is not None):
                            f.write(data_str + '\n);\n\nCREATE TABLE ' + str(self._ws.cell(row=i+1, column=1).value).strip() + ' (\n')
                        else:
                            f.write(data_str + '\n);\n\nCREATE TABLE table_name (\n')
                    else:
                        f.write(data_str[:-1] + ',\n')
            else:
                f.write(');')

                
        print('Complete')


#Main Method
def main():
    ExtractExcelUpdate()

if __name__=='__main__':    
    main()
