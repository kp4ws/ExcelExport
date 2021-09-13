import openpyxl
from openpyxl import load_workbook
from datetime import datetime

class ExtractExcel:

    #Constructor
    def __init__(self):

        #Get user input
        print('Enter absolute path for Workbook "C:\\directory name\\<file name>.xlsx"')
        self._wb_name = input()
        #self._wb = load_workbook('.\\'+self._wb_name+'.xlsx', data_only=True)
        self._wb = load_workbook(self._wb_name, data_only=True)
        self._ws = self._wb.active

        print('Enter absolute path for Export Location "C:\\directory name\\<file name>.sql"')
        self._directory = input()
        
        print('Rows:')
        self._rows = int(input())
        print('Columns:')
        self._columns = int(input())

        #column to start at
        print('Enter column start')
        self._col_start = int(input())

    	#Primary method for exporting data
        self.exportData()

    #Export data from excel to sql file
    def exportData(self):
        with open (self._directory, 'w') as f:
            for i in range(2, self._rows+1):
                f.write('INSERT INTO table_name VALUES(');
                data_str = ''

                #modify below number depending on where you want to start from
                #
                #
                #***********Currently starting at value 2********
                #***********BOTH numbers need to start at same value********
                #
                for j in range(self._col_start, self._columns + self._col_start):
                    
                    if(self._ws.cell(row=i,column=j).value is None or self._ws.cell(row=i,column=j).value == 'NULL'):
                        data_str += 'NULL,'

                    elif (isinstance(self._ws.cell(row=i,column=j).value, int)):
                        data_str += str(self._ws.cell(row=i, column=j).value) + ","

                    else:
                        data_str += "'"+ str(self._ws.cell(row=i, column=j).value) + "',"

                else:
                    f.write(data_str[:-1] + ')\n')
                
        print('Complete')


#Main Method
def main():
    ExtractExcel()

if __name__=='__main__':    
    main()
